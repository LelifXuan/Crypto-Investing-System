from __future__ import annotations

# ruff: noqa: E501
import csv
import json
from pathlib import Path
from typing import Any, Mapping

from app.core.paths import app_paths

DEFAULT_GOLDHUB_DIR = app_paths.resource_root / "data" / "goldhub"

CATEGORY_PATTERNS = {
    "central_bank": ("central_bank", "reserve"),
    "gold_etf": ("gold_etf",),
    "supply_demand": ("supply_demand",),
    "mine_production": ("mine_production",),
    "aisc": ("aisc",),
    "futures_oi": ("futures_oi",),
}


def _float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _first_float(payload: Mapping[str, Any], *keys: str) -> float | None:
    for key in keys:
        value = _float(payload.get(key))
        if value is not None:
            return value
    return None


def _first_value(payload: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        value = payload.get(key)
        if value not in (None, ""):
            return value
    return None


def _read_json(path: Path) -> dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if isinstance(data, list):
        return dict(data[-1]) if data and isinstance(data[-1], Mapping) else {}
    return dict(data) if isinstance(data, Mapping) else {}


def _read_csv(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.DictReader(fh))
    except OSError:
        return {}
    return dict(rows[-1]) if rows else {}


def _read_xlsx(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except OSError:
        return {}
    if len(rows) < 2:
        return {}
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_row = rows[-1]
    return {header: value for header, value in zip(headers, data_row, strict=False) if header}


def _read_payload(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _read_json(path)
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return {}


def _normalize_from_categories(raw: Mapping[str, Mapping[str, Any]], missing_categories: list[str]) -> dict[str, Any]:
    central = raw.get("central_bank", {})
    etf = raw.get("gold_etf", {})
    supply = {**raw.get("supply_demand", {}), **raw.get("mine_production", {}), **raw.get("aisc", {})}
    derivatives = raw.get("futures_oi", {})
    updated_at = _first_value(
        {**central, **etf, **supply, **derivatives},
        "updated_at",
        "date",
        "period",
        "latest_date",
    )
    snapshot = {
        "source": "world_gold_council_goldhub_local",
        "updated_at": str(updated_at) if updated_at is not None else None,
        "central_bank_net_purchase_tonnes_12m": _first_float(central, "central_bank_net_purchase_tonnes_12m", "net_purchase_tonnes_12m", "net_buying_12m", "12m_net_purchase"),
        "central_bank_net_purchase_tonnes_3m": _first_float(central, "central_bank_net_purchase_tonnes_3m", "net_purchase_tonnes_3m", "net_buying_3m", "3m_net_purchase"),
        "gold_etf_flow_tonnes_1m": _first_float(etf, "gold_etf_flow_tonnes_1m", "etf_flow_tonnes_1m", "flow_tonnes_1m"),
        "mine_production_yoy": _first_float(supply, "mine_production_yoy", "mine_yoy", "production_yoy"),
        "recycling_yoy": _first_float(supply, "recycling_yoy", "recycle_yoy"),
        "aisc_yoy": _first_float(supply, "aisc_yoy", "all_in_sustaining_cost_yoy"),
        "supply_demand_balance_tonnes": _first_float(supply, "supply_demand_balance_tonnes", "balance_tonnes", "market_balance_tonnes"),
        "futures_oi_change_4w": _first_float(derivatives, "futures_oi_change_4w", "oi_change_4w", "open_interest_change_4w"),
        "futures_volume_zscore": _first_float(derivatives, "futures_volume_zscore", "volume_zscore"),
        "cot_net_spec_percentile": _first_float(derivatives, "cot_net_spec_percentile", "cot_percentile"),
    }
    snapshot.update(
        {
            "central_bank": {
                "central_bank_net_purchase_tonnes_12m": snapshot["central_bank_net_purchase_tonnes_12m"],
                "central_bank_net_purchase_tonnes_3m": snapshot["central_bank_net_purchase_tonnes_3m"],
            },
            "supply": {
                "mine_production_yoy": snapshot["mine_production_yoy"],
                "recycling_yoy": snapshot["recycling_yoy"],
                "aisc_yoy": snapshot["aisc_yoy"],
                "supply_demand_balance_tonnes": snapshot["supply_demand_balance_tonnes"],
            },
            "investment_flow": {
                "gold_etf_flow_tonnes_1m": snapshot["gold_etf_flow_tonnes_1m"],
            },
            "derivatives": {
                "futures_oi_change_4w": snapshot["futures_oi_change_4w"],
                "futures_volume_zscore": snapshot["futures_volume_zscore"],
                "cot_net_spec_percentile": snapshot["cot_net_spec_percentile"],
            },
            "performance_metrics": {},
            "data_quality": {
                "file_available": len(missing_categories) < len(CATEGORY_PATTERNS),
                "missing_categories": missing_categories,
            },
        }
    )
    return snapshot


def normalize_goldhub_snapshot(raw: Mapping[str, Any] | None = None) -> dict[str, Any]:
    raw = dict(raw or {})
    categories = {
        "central_bank": dict(raw.get("central_bank") or {}),
        "gold_etf": dict(raw.get("investment_flow") or {}),
        "supply_demand": dict(raw.get("supply") or {}),
        "futures_oi": dict(raw.get("derivatives") or {}),
    }
    for key, value in raw.items():
        if key not in {"central_bank", "investment_flow", "supply", "derivatives"}:
            categories.setdefault("flat", {})[key] = value
    if "flat" in categories:
        for category in ("central_bank", "gold_etf", "supply_demand", "futures_oi"):
            categories[category] = {**categories.get("flat", {}), **categories.get(category, {})}
    missing = [key for key in CATEGORY_PATTERNS if not categories.get(key)]
    snapshot = _normalize_from_categories(categories, missing)
    snapshot["source"] = raw.get("source", snapshot["source"])
    snapshot["updated_at"] = raw.get("updated_at", snapshot.get("updated_at"))
    snapshot["data_quality"] = {**snapshot["data_quality"], **dict(raw.get("data_quality") or {})}
    return snapshot


class GoldhubDataService:
    def __init__(self, data_dir: Path | None = None) -> None:
        self.data_dir = data_dir or DEFAULT_GOLDHUB_DIR

    def load_snapshot(self) -> dict[str, Any]:
        raw_by_category = self._read_semantic_payloads()
        missing = [key for key in CATEGORY_PATTERNS if key not in raw_by_category]
        snapshot = _normalize_from_categories(raw_by_category, missing)
        snapshot["data_quality"] = {
            **snapshot.get("data_quality", {}),
            "data_dir": str(self.data_dir),
        }
        return snapshot

    def _read_semantic_payloads(self) -> dict[str, dict[str, Any]]:
        if not self.data_dir.exists():
            return {}
        files = [item for item in self.data_dir.iterdir() if item.is_file() and item.suffix.lower() in {".json", ".csv", ".xlsx", ".xlsm"}]
        payloads: dict[str, dict[str, Any]] = {}
        for category, patterns in CATEGORY_PATTERNS.items():
            matches = [file for file in files if any(file.name.lower().startswith(pattern) for pattern in patterns)]
            if not matches:
                continue
            latest = max(matches, key=lambda item: item.stat().st_mtime)
            payload = _read_payload(latest)
            if payload:
                payloads[category] = payload
        return payloads
